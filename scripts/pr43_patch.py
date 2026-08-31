from pathlib import Path


def replace_once(text: str, old: str, new: str, label: str) -> str:
    if old not in text:
        raise SystemExit(f"missing patch anchor: {label}")
    return text.replace(old, new, 1)

bulk_path = Path("apps/desktop/src-tauri/src/workspace/bulk_model.rs")
bulk = bulk_path.read_text()
bulk = replace_once(
    bulk,
    "        default_value: Option<String>,\n        flow_direction: Option<FlowDirection>,\n",
    "        default_value: Option<String>,\n        flow_direction: Option<FlowDirection>,\n        is_conjugated: Option<bool>,\n",
    "bulk update fields declaration",
)
bulk = replace_once(
    bulk,
    "                    default_value,\n                    flow_direction,\n                } => {",
    "                    default_value,\n                    flow_direction,\n                    is_conjugated,\n                } => {",
    "bulk update fields destructure",
)
bulk = replace_once(
    bulk,
    "                    if requirement_id.is_some() || requirement_text.is_some() {",
    '''                    if let Some(is_conjugated) = is_conjugated {
                        let kind = project
                            .element(id)
                            .map_err(|cause| {
                                error("SEMANTIC_VALIDATION", Some(index), cause.to_string())
                            })?
                            .kind
                            .clone();
                        if !matches!(kind, ElementKind::ProxyPort | ElementKind::FullPort) {
                            return Err(error(
                                "SEMANTIC_VALIDATION",
                                Some(index),
                                "Conjugated mapping is valid only for ProxyPort or FullPort",
                            ));
                        }
                        project
                            .element_mut(id)
                            .map_err(|cause| {
                                error("SEMANTIC_VALIDATION", Some(index), cause.to_string())
                            })?
                            .is_conjugated = *is_conjugated;
                    }
                    if requirement_id.is_some() || requirement_text.is_some() {''',
    "bulk conjugation mutation",
)
# Existing bulk-model tests construct this variant with explicit None values.
bulk = bulk.replace(
    "                    flow_direction: None,\n",
    "                    flow_direction: None,\n                    is_conjugated: None,\n",
)
bulk_path.write_text(bulk)

sheet_path = Path("apps/desktop/src-tauri/src/workspace/spreadsheet_import.rs")
sheet = sheet_path.read_text()
sheet = replace_once(
    sheet,
    "    DefaultValue,\n    FlowDirection,\n    ExternalId,\n",
    "    DefaultValue,\n    FlowDirection,\n    Conjugated,\n    ExternalId,\n",
    "semantic Conjugated vocabulary",
)
sheet = replace_once(
    sheet,
    "            | ElementKind::ConstraintProperty\n            | ElementKind::ConstraintParameter\n    )\n}\n\nfn is_namespace_kind",
    "            | ElementKind::ConstraintProperty\n            | ElementKind::ConstraintParameter\n            | ElementKind::ProxyPort\n            | ElementKind::FullPort\n    )\n}\n\nfn is_namespace_kind",
    "supported kinds ports",
)
sheet = replace_once(
    sheet,
    "            | ElementKind::ConstraintProperty\n            | ElementKind::ConstraintParameter\n    )\n}\n\nfn supported_relationship_kind",
    "            | ElementKind::ConstraintProperty\n            | ElementKind::ConstraintParameter\n            | ElementKind::ProxyPort\n            | ElementKind::FullPort\n    )\n}\n\nfn supported_relationship_kind",
    "feature kinds ports",
)
sheet = replace_once(
    sheet,
    "            SpreadsheetSemanticProperty::FlowDirection,\n            SpreadsheetSemanticProperty::RequirementId,",
    "            SpreadsheetSemanticProperty::FlowDirection,\n            SpreadsheetSemanticProperty::Conjugated,\n            SpreadsheetSemanticProperty::RequirementId,",
    "relationship feature-only Conjugated",
)
sheet = replace_once(
    sheet,
    "            || has_property(SpreadsheetSemanticProperty::DefaultValue)\n            || has_property(SpreadsheetSemanticProperty::FlowDirection))",
    "            || has_property(SpreadsheetSemanticProperty::DefaultValue)\n            || has_property(SpreadsheetSemanticProperty::FlowDirection)\n            || has_property(SpreadsheetSemanticProperty::Conjugated))",
    "non-feature reserved Conjugated",
)
sheet = replace_once(
    sheet,
    '            "Type/Multiplicity/Default Value/Flow Direction mappings are reserved for PR39 owned features",',
    '            "Type/Multiplicity/Default Value/Flow Direction/Conjugated mappings are reserved for owned features",',
    "reserved feature diagnostic",
)
sheet = replace_once(
    sheet,
    "    let target = project.element(map.target_scope).map_err(|_| {",
    '''    if !matches!(map.element_kind, ElementKind::ProxyPort | ElementKind::FullPort)
        && has_property(SpreadsheetSemanticProperty::Conjugated)
    {
        return Err(diagnostic(
            Some(map),
            None,
            mapped_column_name(map, SpreadsheetSemanticProperty::Conjugated),
            Some(SpreadsheetSemanticProperty::Conjugated),
            None,
            "SEMANTIC_PROPERTY_INVALID",
            "Conjugated can be mapped only for ProxyPort or FullPort",
        ));
    }

    let target = project.element(map.target_scope).map_err(|_| {''',
    "port conjugation map validation",
)

# Owner-qualified fallback identity for ports: a name alone is never sufficient across owners.
sheet = replace_once(
    sheet,
    "fn find_existing<'a>(\n    map: &SpreadsheetImportMap,\n    project: &'a Project,\n    values: &BTreeMap<SpreadsheetSemanticProperty, String>,\n) -> Result<Option<&'a Element>, SpreadsheetImportDiagnostic> {",
    "fn find_existing<'a>(\n    map: &SpreadsheetImportMap,\n    project: &'a Project,\n    values: &BTreeMap<SpreadsheetSemanticProperty, String>,\n    resolved_owner: Option<&ResolvedOwner>,\n) -> Result<Option<&'a Element>, SpreadsheetImportDiagnostic> {",
    "find_existing owner argument",
)
sheet = replace_once(
    sheet,
    "        .filter(|element| {\n            existing_match_in_scope(project, element, map.target_scope, map.search_scope)\n        })\n        .filter(|element| match map.identification_property {",
    '''        .filter(|element| {
            existing_match_in_scope(project, element, map.target_scope, map.search_scope)
        })
        .filter(|element| {
            if map.identification_property == SpreadsheetIdentificationProperty::Name
                && matches!(map.element_kind, ElementKind::ProxyPort | ElementKind::FullPort)
            {
                match resolved_owner.map(|owner| &owner.reference) {
                    Some(BuildReference::Existing(owner_id)) => element.owner_id == Some(*owner_id),
                    Some(BuildReference::External(_)) | None => false,
                }
            } else {
                true
            }
        })
        .filter(|element| match map.identification_property {''',
    "owner-qualified port fallback",
)

sheet = replace_once(
    sheet,
    "fn parse_aggregation(\n",
    '''fn parse_conjugated(
    map: &SpreadsheetImportMap,
    row: usize,
    value: &str,
) -> Result<bool, SpreadsheetImportDiagnostic> {
    match value.trim().to_ascii_lowercase().as_str() {
        "true" | "yes" | "1" => Ok(true),
        "false" | "no" | "0" => Ok(false),
        _ => Err(diagnostic(
            Some(map),
            Some(row),
            mapped_column_name(map, SpreadsheetSemanticProperty::Conjugated),
            Some(SpreadsheetSemanticProperty::Conjugated),
            None,
            "CONJUGATED_INVALID",
            format!(
                "conjugation '{}' must be true/false, yes/no, or 1/0",
                value
            ),
        )),
    }
}

fn parse_aggregation(
''',
    "conjugation parser",
)

sheet = replace_once(
    sheet,
    "    multiplicity: Option<Multiplicity>,\n    flow_direction: Option<FlowDirection>,\n    values: &BTreeMap<SpreadsheetSemanticProperty, String>,",
    "    multiplicity: Option<Multiplicity>,\n    flow_direction: Option<FlowDirection>,\n    is_conjugated: Option<bool>,\n    values: &BTreeMap<SpreadsheetSemanticProperty, String>,",
    "mapped field changes signature",
)
sheet = replace_once(
    sheet,
    "        || flow_direction.is_some_and(|value| element.flow_direction != Some(value))\n        || default_changed",
    "        || flow_direction.is_some_and(|value| element.flow_direction != Some(value))\n        || is_conjugated.is_some_and(|value| element.is_conjugated != value)\n        || default_changed",
    "conjugation change detection",
)
sheet = replace_once(
    sheet,
    "            default_value,\n            flow_direction,\n        },",
    "            default_value,\n            flow_direction,\n            is_conjugated,\n        },",
    "existing port update operation",
)

sheet = replace_once(
    sheet,
    "            let existing = match find_existing(map, project, &values) {",
    '''            let is_conjugated = if matches!(
                map.element_kind,
                ElementKind::ProxyPort | ElementKind::FullPort
            ) {
                match non_empty_value(&values, SpreadsheetSemanticProperty::Conjugated) {
                    Some(value) => match parse_conjugated(map, row.row_number, value) {
                        Ok(value) => Some(value),
                        Err(error) => {
                            block_row(error);
                            continue;
                        }
                    },
                    None => None,
                }
            } else {
                None
            };

            let existing = match find_existing(map, project, &values, Some(&owner)) {''',
    "parse conjugation and owner-qualified find",
)
sheet = replace_once(
    sheet,
    "                    multiplicity,\n                    flow_direction,\n                    &values,",
    "                    multiplicity,\n                    flow_direction,\n                    is_conjugated,\n                    &values,",
    "mapped field changes call",
)
sheet = replace_once(
    sheet,
    "                || default_value.is_some()\n                || flow_direction.is_some()",
    "                || default_value.is_some()\n                || flow_direction.is_some()\n                || is_conjugated.is_some()",
    "create scalar update condition",
)
sheet = replace_once(
    sheet,
    "                    default_value,\n                    flow_direction,\n                });",
    "                    default_value,\n                    flow_direction,\n                    is_conjugated,\n                });",
    "create port update operation",
)

# Keep any existing constructors complete after the bulk-model field addition.
sheet = sheet.replace(
    "            flow_direction: None,\n        },",
    "            flow_direction: None,\n            is_conjugated: None,\n        },",
)

if "mod pr43_tests;" not in sheet:
    sheet += "\n#[cfg(test)]\nmod pr43_tests;\n"
sheet_path.write_text(sheet)

portable_path = Path("apps/desktop/src-tauri/src/workspace/portable_interchange.rs")
portable = portable_path.read_text()
if "mod pr43_port_tests;" not in portable:
    portable += "\n#[cfg(test)]\nmod pr43_port_tests;\n"
portable_path.write_text(portable)

# Generate the compact CATIA-style workbook fixture.
from openpyxl import Workbook
fixture = Path("apps/desktop/src-tauri/tests/fixtures/pr43_ports.xlsx")
fixture.parent.mkdir(parents=True, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = "Interface Definitions"
ws.append(["Interface ID", "Interface Name"])
ws.append(["IF-CMD", "CommandInterface"])
ws.append(["IF-TEL", "TelemetryInterface"])
ws = wb.create_sheet("Service Types")
ws.append(["Type ID", "Type Name"])
ws.append(["DT-SVC", "ServiceAssembly"])
ws = wb.create_sheet("Components")
ws.append(["Component ID", "Component Name"])
ws.append(["BLK-VEH", "Vehicle"])
ws.append(["BLK-CTRL", "Controller"])
ws.append(["BLK-PWR", "PowerUnit"])
ws = wb.create_sheet("Component Interfaces")
ws.append(["Port Identifier", "Owning Component", "Interface Name", "Interface Type", "Cardinality", "Conjugated", "Description", "Access"])
ws.append(["PORT-CMD", "BLK-VEH", "command", "IF-CMD", "1", "false", "Vehicle command port", "Public"])
ws.append(["PORT-TEL", "BLK-CTRL", "telemetry", "IF-TEL", "0..1", "0", "Telemetry port", "Public"])
ws.append(["PORT-PEER", "BLK-CTRL", "commandPeer", "IF-CMD", "1", "yes", "Conjugated command peer", "Private"])
ws = wb.create_sheet("Service Ports")
ws.append(["Port ID", "Component", "Port Name", "Port Type", "Multiplicity", "Conjugated", "Description", "Visibility"])
ws.append(["PORT-SVC", "BLK-PWR", "service", "DT-SVC", "1", "false", "Full service port", "Private"])
wb.save(fixture)
